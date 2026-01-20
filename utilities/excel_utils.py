import openpyxl
from openpyxl.styles import PatternFill

def get_row_count(file_path, sheet_name):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    return sheet.max_row

def get_column_count(file_path, sheet_name):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    return sheet.max_column

def read_cell_data(file_path, sheet_name, row, column):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    return sheet.cell(row=row, column=column).value

def write_cell_data(file_path, sheet_name, row, column, data):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    sheet.cell(row=row, column=column).value = data
    workbook.save(file_path)
